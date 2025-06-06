import sys
import openpyxl
from openpyxl.styles import Border, Font, Alignment, Side

linecolor = '000000'
font_style = 'Arial'
font_size = 10

def generate_xlsx(si_style, SI_workbook, file, basis_set, charge, multiplicity, total_energy, jobtype, imaginary_freqs, coords, state_blocks, energies, wavelengths, f_osc):
	
	if si_style == 4:
		print('Generating Full .xlsx file ...')
		# Generate Excel file and format cells
		SI_worksheet = SI_workbook.create_sheet(title=file.strip()[:-4])
		mediumBorder = Side(border_style = 'medium', color = linecolor)
		thinBorder = Side(border_style = 'thin', color = linecolor)
		dashedBorder = Side(border_style = 'dashed', color = linecolor)

		SI_worksheet.column_dimensions['A'].width = 7.25
		SI_worksheet.column_dimensions['B'].width = 10.2
		SI_worksheet.column_dimensions['C'].width = 10.2
		SI_worksheet.column_dimensions['D'].width = 10.2
		SI_worksheet.column_dimensions['E'].width = 7.25
		SI_worksheet.column_dimensions['F'].width = 10.2
		SI_worksheet.column_dimensions['G'].width = 10.2
		SI_worksheet.column_dimensions['H'].width = 10.2
		for w in range(1, 5000):
			SI_worksheet.row_dimensions[w].height = 16.0

		cellRange3 = SI_worksheet['A1:H1']
		for cellNum3 in cellRange3:
			for cellNum4 in cellNum3:
				cellNum4.border = Border(bottom = mediumBorder)

		cellRange5 = SI_worksheet['A9:H9']
		for cellNum5 in cellRange5:
			for cellNum6 in cellNum5:
				cellNum6.border = Border(bottom = thinBorder)

		cellRange = SI_worksheet['A2:H9']
		for cellNum in cellRange:
			for cellNum1 in cellNum:
				cellNum1.font = Font(name = font_style, size = font_size)

		SI_worksheet.merge_cells('A1:H1')
		SI_worksheet['A1'] = file
		SI_worksheet['A1'].font = Font(name = font_style, size = 10.5, bold = True)
		SI_worksheet['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

		SI_worksheet.merge_cells('A2:C9')
		SI_worksheet['A2'].font = Font(name = font_style, size = 10.5, color = '929292')
		SI_worksheet['A2'].alignment = Alignment(horizontal = 'center', vertical = 'center')
		SI_worksheet['A2'] = 'Insert molecular geometry here.'

		# Hier route section einführen
		SI_worksheet.merge_cells('D2:H2')
		SI_worksheet['D2'] = f'Basis set: {basis_set}'
		SI_worksheet['D2'].font = Font(name = 'Courier', size = font_size)

		SI_worksheet.merge_cells('D3:H3')
		SI_worksheet['D3'] = f'Charge = {charge}, Multiplicity = {multiplicity}'

		SI_worksheet.merge_cells('D4:H4')
		SI_worksheet['D4'] = f'Electronic Energy = {total_energy} Hartree'

		SI_worksheet.merge_cells('D5:H5')
		if jobtype == 'opt+freq' or jobtype == 'freq':
			if len(imaginary_freqs) == 0:
				SI_worksheet['D5'] = f'Number of imaginary frequencies = 0'
			else:
				SI_worksheet['D5'] = f'Number of imaginary frequencies = {len(imaginary_freqs)}, v_i = {', '.join(imaginary_freqs)} cm-1'

		SI_worksheet.merge_cells('D6:H6')
		SI_worksheet.merge_cells('D7:H7')
		SI_worksheet.merge_cells('D8:H8')
		SI_worksheet.merge_cells('D9:H9')

		'''
		if jobType == 'opt+freq' or jobType == 'freq':
			SI_worksheet['D6'] = f'Sum of electronic and zero-point Energies = {freqZPE} Hartree'
			SI_worksheet['D7'] = f'Sum of electronic and thermal Energies = {freqThr} Hartree'
			SI_worksheet['D8'] = f'Sum of electronic and thermal Enthalpies = {freqH} Hartree'
			SI_worksheet['D9'] = f'Sum of electronic and thermal Free Energies = {freqFE} Hartree'
		'''

		if len(coords) > 0:
			SI_worksheet.merge_cells('A10:A11')
			SI_worksheet.merge_cells('B10:D10')
			SI_worksheet.merge_cells('E10:E11')
			SI_worksheet.merge_cells('F10:H10')

			SI_worksheet['A10'] = 'Atoms'
			SI_worksheet['A10'].font = Font(name = font_style, size = font_size, bold = True)
			SI_worksheet['A10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
			SI_worksheet['A11'].border = Border(bottom = thinBorder)

			SI_worksheet['B10'] = 'Cartesian Coordinates'
			SI_worksheet['B10'].font = Font(name = font_style, size = font_size, bold = True)
			SI_worksheet['B10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			SI_worksheet['B11'] = 'X'
			SI_worksheet['B11'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['B11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['B11'].border = Border(top = dashedBorder, bottom = thinBorder)
			SI_worksheet['C11'] = 'Y'
			SI_worksheet['C11'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['C11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['C11'].border = Border(top = dashedBorder, bottom = thinBorder)
			SI_worksheet['D11'] = 'Z'
			SI_worksheet['D11'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['D11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['D11'].border = Border(top = dashedBorder, bottom = thinBorder)

			SI_worksheet['E10'] = 'Atoms'
			SI_worksheet['E10'].font = Font(name = font_style, size = font_size, bold = True)
			SI_worksheet['E10'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
			SI_worksheet['E11'].border = Border(bottom = thinBorder)

			SI_worksheet['F10'] = 'Cartesian Coordinates'
			SI_worksheet['F10'].font = Font(name = font_style, size = font_size, bold = True)
			SI_worksheet['F10'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			SI_worksheet['F11'] = 'X'
			SI_worksheet['F11'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['F11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['F11'].border = Border(top = dashedBorder, bottom = thinBorder)
			SI_worksheet['G11'] = 'Y'
			SI_worksheet['G11'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['G11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['G11'].border = Border(top = dashedBorder, bottom = thinBorder)
			SI_worksheet['H11'] = 'Z'
			SI_worksheet['H11'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['H11'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['H11'].border = Border(top = dashedBorder, bottom = thinBorder)

			# Fill cells with coordinates
			excel_coords = []
			excel_coordss = []

			if len(coords) % 2 == 0:
				excel_coordsLineNumber = int(len(coords) / 2)
				for r in range(0, len(coords), 2):
					excel_coords = []
					excel_coords.append(coords[r].split()[0])
					excel_coords.append(coords[r].split()[1])
					excel_coords.append(coords[r].split()[2])
					excel_coords.append(coords[r].split()[3])
					excel_coords.append(coords[r + 1].split()[0])
					excel_coords.append(coords[r + 1].split()[1])
					excel_coords.append(coords[r + 1].split()[2])
					excel_coords.append(coords[r + 1].split()[3])
					excel_coordss.append(excel_coords)

			elif len(coords) % 2 == 1:
				excel_coordsLineNumber = int((len(coords) + 1) / 2)
				for r in range(0, len(coords) - 1, 2):
					excel_coords = []
					excel_coords.append(coords[r].split()[0])
					excel_coords.append(coords[r].split()[1])
					excel_coords.append(coords[r].split()[2])
					excel_coords.append(coords[r].split()[3])
					excel_coords.append(coords[r + 1].split()[0])
					excel_coords.append(coords[r + 1].split()[1])
					excel_coords.append(coords[r + 1].split()[2])
					excel_coords.append(coords[r + 1].split()[3])
					excel_coordss.append(excel_coords)
				excel_coords = []
				excel_coords.append(coords[-1].split()[0])
				excel_coords.append(coords[-1].split()[1])
				excel_coords.append(coords[-1].split()[2])
				excel_coords.append(coords[-1].split()[3])
				excel_coordss.append(excel_coords)

			for line in excel_coordss:
				SI_worksheet.append(line)

			cellRange2 = SI_worksheet['A12:H4001']
			for cellNum in cellRange2:
				for cellNum1 in cellNum:
					cellNum1.font = Font(name = font_style, size = font_size)
					cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

			lastLine = excel_coordsLineNumber + 11

			cellRange7 = SI_worksheet[f'A{lastLine}:H{lastLine}']
			for cellNum7 in cellRange7:
				for cellNum8 in cellNum7:
					cellNum8.border = Border(bottom = mediumBorder)
			cellRange9 = SI_worksheet[f'E10:E{lastLine}']
			for cellNum9 in cellRange9:
				for cellNumfont_size in cellNum9:
					cellNumfont_size.border = Border(left = dashedBorder)
			SI_worksheet['E11'].border = Border(left = dashedBorder, bottom = thinBorder)
			SI_worksheet[f'E{lastLine}'].border = Border(left = dashedBorder, bottom = mediumBorder)

	elif si_style == 5:
		print('Generating Simple .xlsx file ...')

		# Generate Excel file and format cells
		SI_worksheet = SI_workbook.create_sheet(title=file.strip()[:-4])
		mediumBorder = Side(border_style = 'medium', color = linecolor)
		thinBorder = Side(border_style = 'thin', color = linecolor)
		dashedBorder = Side(border_style = 'dashed', color = linecolor)

		SI_worksheet.column_dimensions['A'].width = 8.25
		SI_worksheet.column_dimensions['B'].width = 12.25
		SI_worksheet.column_dimensions['C'].width = 12.25
		SI_worksheet.column_dimensions['D'].width = 12.25

		cellRange = SI_worksheet['A2:D4']
		for cellNum in cellRange:
			for cellNum1 in cellNum:
				cellNum1.font = Font(name = font_style, size = 10)

		for w in range(1, 8000):
			SI_worksheet.row_dimensions[w].height = 16.0

		cellRange3 = SI_worksheet['A1:D1']
		for cellNum3 in cellRange3:
			for cellNum4 in cellNum3:
				cellNum4.border = Border(bottom = mediumBorder)

		cellRange5 = SI_worksheet['A4:D4']
		for cellNum5 in cellRange5:
			for cellNum6 in cellNum5:
				cellNum6.border = Border(bottom = thinBorder)

		SI_worksheet.merge_cells('A1:D1')
		SI_worksheet['A1'] = file
		SI_worksheet['A1'].font = Font(name = font_style, size = font_size, bold = True)
		SI_worksheet['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

		SI_worksheet.merge_cells('A2:D2')
		SI_worksheet['A2'] = f'Basis set: {basis_set}'
		SI_worksheet['A2'].font = Font(name = 'Courier', size = font_size)

		SI_worksheet.merge_cells('A3:D3')
		SI_worksheet['A3'] = f'Charge = {charge}, Multiplicity = {multiplicity}'

		SI_worksheet.merge_cells('A4:D4')
		SI_worksheet['A4'] = f'Electronic Energy = {total_energy} Hartree'

		if len(coords) > 0:

			SI_worksheet.merge_cells('A5:A6')
			SI_worksheet.merge_cells('B5:D5')

			SI_worksheet['A5'] = 'Atoms'
			SI_worksheet['A5'].font = Font(name = font_style, size = font_size, bold = True)
			SI_worksheet['A5'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
			SI_worksheet['A6'].border = Border(bottom = thinBorder)

			SI_worksheet['B5'] = 'Cartesian Coordinates'
			SI_worksheet['B5'].font = Font(name = font_style, size = font_size, bold = True)
			SI_worksheet['B5'].alignment = Alignment(horizontal = 'center', vertical = 'center')

			SI_worksheet['B6'] = 'X'
			SI_worksheet['B6'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['B6'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['B6'].border = Border(top = dashedBorder, bottom = thinBorder)
			SI_worksheet['C6'] = 'Y'
			SI_worksheet['C6'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['C6'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['C6'].border = Border(top = dashedBorder, bottom = thinBorder)
			SI_worksheet['D6'] = 'Z'
			SI_worksheet['D6'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
			SI_worksheet['D6'].alignment = Alignment(horizontal = 'center', vertical = 'center')
			SI_worksheet['D6'].border = Border(top = dashedBorder, bottom = thinBorder)

			excel_coordss = []
			for s in range(len(coords)):
				excel_coords = []
				excel_coords.append(coords[s].split()[0])
				excel_coords.append(coords[s].split()[1])
				excel_coords.append(coords[s].split()[2])
				excel_coords.append(coords[s].split()[3])
				excel_coordss.append(excel_coords)

			for line in excel_coordss:
				SI_worksheet.append(line)
			
			cellRange1 = SI_worksheet['A7:D8007']
			for cellNum in cellRange1:
				for cellNum1 in cellNum:
					cellNum1.font = Font(name = font_style, size = 10)
					cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

			lastLine = len(coords) + 6

			cellRange7 = SI_worksheet[f'A{lastLine}:D{lastLine}']
			for cellNum7 in cellRange7:
				for cellNum8 in cellNum7:
					cellNum8.border = Border(bottom = mediumBorder)

	elif si_style == 6:
		if len(coords) == 0:
			print('No coordinates found in output file. Exiting the program...')
			sys.exit()

		print('Generating coordinates .xlsx file ...')

		# Generate Excel file and format cells
		SI_worksheet = SI_workbook.create_sheet(title=file.strip()[:-4])
		mediumBorder = Side(border_style = 'medium', color = linecolor)
		thinBorder = Side(border_style = 'thin', color = linecolor)
		dashedBorder = Side(border_style = 'dashed', color = linecolor)

		SI_worksheet.column_dimensions['A'].width = 7.25
		SI_worksheet.column_dimensions['B'].width = 11.25
		SI_worksheet.column_dimensions['C'].width = 11.25
		SI_worksheet.column_dimensions['D'].width = 11.25

		for w in range(1, 8000):
			SI_worksheet.row_dimensions[w].height = 16.0

		cellRange3 = SI_worksheet['A1:D1']
		for cellNum3 in cellRange3:
			for cellNum4 in cellNum3:
				cellNum4.border = Border(bottom = mediumBorder)

		SI_worksheet.merge_cells('A1:D1')
		SI_worksheet['A1'] = file
		SI_worksheet['A1'].font = Font(name = font_style, size = font_size, bold = True)
		SI_worksheet['A1'].alignment = Alignment(horizontal = 'center', vertical = 'center')

		SI_worksheet.merge_cells('A2:A3')
		SI_worksheet.merge_cells('B2:D2')

		SI_worksheet['A2'] = 'Atoms'
		SI_worksheet['A2'].font = Font(name = font_style, size = font_size, bold = True)
		SI_worksheet['A2'].alignment = Alignment(horizontal = 'center', vertical = 'bottom')
		SI_worksheet['A3'].border = Border(bottom = thinBorder)

		SI_worksheet['B2'] = 'Cartesian Coordinates'
		SI_worksheet['B2'].font = Font(name = font_style, size = font_size, bold = True)
		SI_worksheet['B2'].alignment = Alignment(horizontal = 'center', vertical = 'center')

		SI_worksheet['B3'] = 'X'
		SI_worksheet['B3'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
		SI_worksheet['B3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
		SI_worksheet['B3'].border = Border(top = dashedBorder, bottom = thinBorder)
		SI_worksheet['C3'] = 'Y'
		SI_worksheet['C3'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
		SI_worksheet['C3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
		SI_worksheet['C3'].border = Border(top = dashedBorder, bottom = thinBorder)
		SI_worksheet['D3'] = 'Z'
		SI_worksheet['D3'].font = Font(name = font_style, size = font_size, bold = True, italic = True)
		SI_worksheet['D3'].alignment = Alignment(horizontal = 'center', vertical = 'center')
		SI_worksheet['D3'].border = Border(top = dashedBorder, bottom = thinBorder)

		excel_coordss = []
		for s in range(len(coords)):
			excel_coords = []
			excel_coords.append(coords[s].split()[0])
			excel_coords.append(coords[s].split()[1])
			excel_coords.append(coords[s].split()[2])
			excel_coords.append(coords[s].split()[3])
			excel_coordss.append(excel_coords)

		for line in excel_coordss:
			SI_worksheet.append(line)
		
		cellRange1 = SI_worksheet['A4:D8005']
		for cellNum in cellRange1:
			for cellNum1 in cellNum:
				cellNum1.font = Font(name = font_style, size = font_size)
				cellNum1.alignment = Alignment(horizontal = 'center', vertical = 'center')

		lastLine = len(coords) + 3

		cellRange7 = SI_worksheet[f'A{lastLine}:D{lastLine}']
		for cellNum7 in cellRange7:
			for cellNum8 in cellNum7:
				cellNum8.border = Border(bottom = mediumBorder)
				
	if jobtype == 'tddft':
		write_tddft = input('Write TD-DFT summary ([yes]/no)? ').strip().lower() or ('yes')
		if write_tddft == 'yes':
			print('Writing TD-DFT summary...')
			# Generate Excel file and format cells
			SI_worksheet = SI_workbook.create_sheet(title=file.strip()[:-4] + ' TD-DFT Summary')
			mediumBorder = Side(border_style = 'medium', color = linecolor)
			thinBorder = Side(border_style = 'thin', color = linecolor)
			dashedBorder = Side(border_style = 'dashed', color = linecolor)
			SI_worksheet.column_dimensions['A'].width = 10
			SI_worksheet.column_dimensions['B'].width = 21
			SI_worksheet.column_dimensions['C'].width = 15
			SI_worksheet.column_dimensions['D'].width = 15
			SI_worksheet.column_dimensions['E'].width = 15

			cellRange3 = SI_worksheet['A1:E1']
			for cellNum3 in cellRange3:
				for cellNum4 in cellNum3:
					cellNum4.border = Border(bottom = thinBorder)
					cellNum4.font = Font(name = font_style, size = 10, bold = True)
					cellNum4.alignment = Alignment(horizontal = 'center', vertical = 'center')
					
			SI_worksheet['A1'] = 'State'
			SI_worksheet['B1'] = 'Orbital Contribution'
			SI_worksheet['C1'] = 'Energy (eV)'
			SI_worksheet['D1'] = 'Wavelength (nm)'
			SI_worksheet['E1'] = 'f_osc'

			row = 2

			wrap_center = Alignment(wrap_text=True, vertical='top', horizontal='center')  # Stil für Orbital Contribution
			cell_alignment = Alignment(horizontal='center', vertical='top')
			cell_font = Font(name = font_style, size = font_size)

			def set_cell(ws, coord, value, alignment=cell_alignment, font=cell_font):
				ws[coord] = value
				ws[coord].alignment = alignment
				ws[coord].font = font

			for n in range(len(state_blocks)):
				block_length = len(state_blocks[n])
				start_row = row
				end_row = row + block_length - 1

				orbital_text = '\n'.join([
					f'{sb[0]} → {sb[1]} ({sb[2]:.3f})'
					for sb in state_blocks[n]
				])

				SI_worksheet.merge_cells(f'A{start_row}:A{end_row}')
				SI_worksheet.merge_cells(f'B{start_row}:B{end_row}')
				SI_worksheet.merge_cells(f'C{start_row}:C{end_row}')
				SI_worksheet.merge_cells(f'D{start_row}:D{end_row}')
				SI_worksheet.merge_cells(f'E{start_row}:E{end_row}')

				set_cell(SI_worksheet, f'A{start_row}', n + 1)
				set_cell(SI_worksheet, f'B{start_row}', orbital_text, wrap_center, cell_font)
				set_cell(SI_worksheet, f'C{start_row}', f'{energies[n]:.2f}')
				set_cell(SI_worksheet, f'D{start_row}', f'{wavelengths[n]:.1f}')
				set_cell(SI_worksheet, f'E{start_row}', f'{f_osc[n]}')

				# Bottom border für letzte Zeile setzen
				for col in ['A', 'B', 'C', 'D', 'E']:
					SI_worksheet[f'{col}{end_row}'].border = Border(bottom=thinBorder)

				row = end_row + 1